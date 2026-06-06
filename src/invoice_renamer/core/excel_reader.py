"""Excel读取模块"""
import os
import re
from openpyxl import load_workbook
from typing import List, Dict, Any, Optional, Tuple


class ExcelReader:
    """Excel文件读取器"""

    # 多 sheet 合并时的主表名
    PRIMARY_SHEET = "发票基础信息"
    # 发票号码列名（用于跨 sheet 关联）
    INVOICE_NUMBER_COL = "数电发票号码"
    # 金额列名（用于取最大金额行）
    AMOUNT_COL = "金额"
    # 商品名称原始列名
    GOODS_NAME_COL = "货物或应税劳务名称"
    
    def __init__(self):
        """初始化Excel读取器"""
        pass
    
    def read_excel(self, file_path: str) -> List[Dict[str, Any]]:
        """
        读取Excel文件内容
        
        Args:
            file_path: Excel文件路径
            
        Returns:
            包含所有行数据的字典列表
            
        Raises:
            FileNotFoundError: 文件不存在
            ValueError: 文件格式错误
        """
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"文件不存在: {file_path}")
        
        if not file_path.endswith('.xlsx'):
            raise ValueError("仅支持.xlsx格式的Excel文件")
        
        try:
            workbook = load_workbook(file_path, read_only=True, data_only=True)
            worksheet = workbook.active
            
            # 获取表头
            headers = []
            for cell in worksheet[1]:
                if cell.value is not None:
                    headers.append(str(cell.value))
                else:
                    headers.append(f"Column_{cell.column}")
            
            # 读取数据行
            data = []
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                row_dict = {}
                for col_idx, value in enumerate(row):
                    if col_idx < len(headers):
                        # 将值转换为字符串，保持格式
                        if value is None:
                            row_dict[headers[col_idx]] = ""
                        else:
                            row_dict[headers[col_idx]] = str(value)
                data.append(row_dict)
            
            workbook.close()
            return data
            
        except Exception as e:
            raise ValueError(f"读取Excel文件失败: {str(e)}")
    
    def get_headers(self, file_path: str) -> List[str]:
        """
        获取Excel文件的表头
        
        Args:
            file_path: Excel文件路径
            
        Returns:
            表头列表
        """
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"文件不存在: {file_path}")
        
        if not file_path.endswith('.xlsx'):
            raise ValueError("仅支持.xlsx格式的Excel文件")
        
        try:
            workbook = load_workbook(file_path, read_only=True, data_only=True)
            worksheet = workbook.active
            
            # 获取表头
            headers = []
            for cell in worksheet[1]:
                if cell.value is not None:
                    headers.append(str(cell.value))
                else:
                    headers.append(f"Column_{cell.column}")
            
            workbook.close()
            return headers
            
        except Exception as e:
            raise ValueError(f"读取Excel表头失败: {str(e)}")

    def read_excel_merged(self, file_path: str) -> Tuple[List[Dict[str, Any]], Dict[str, List[str]]]:
        """
        读取Excel文件的所有sheet，按发票号码合并为一行一条发票的数据。

        合并策略：
        - 以"发票基础信息"为主表（1:1）
        - 从"信息汇总表"中取金额最大的明细行的所有独有列（货物或应税劳务名称的值提取商品名）
        - 从"建筑服务"中取独有列
        - 重复列只在主表中显示

        Args:
            file_path: Excel文件路径

        Returns:
            (merged_data, grouped_columns)
            - merged_data: 合并后的字典列表，每行一张发票
            - grouped_columns: 按 sheet 分组的列名字典，如 {"发票基础信息": [...], "信息汇总表": [...], ...}

        Raises:
            FileNotFoundError: 文件不存在
            ValueError: 文件格式错误
        """
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"文件不存在: {file_path}")

        if not file_path.endswith('.xlsx'):
            raise ValueError("仅支持.xlsx格式的Excel文件")

        try:
            workbook = load_workbook(file_path, read_only=True, data_only=True)
            sheet_names = workbook.sheetnames

            # 读取每个 sheet 的 headers 和 data
            sheets_data = {}
            sheets_headers = {}
            for name in sheet_names:
                ws = workbook[name]
                headers = self._read_sheet_headers(ws)
                data = self._read_sheet_data(ws, headers)
                sheets_headers[name] = headers
                sheets_data[name] = data
            workbook.close()

            # 如果只有一个 sheet，退化为原有行为
            if len(sheet_names) == 1:
                only_name = sheet_names[0]
                return sheets_data[only_name], {only_name: sheets_headers[only_name]}

            # 确定主表
            primary_name = self.PRIMARY_SHEET if self.PRIMARY_SHEET in sheet_names else sheet_names[0]
            primary_data = sheets_data[primary_name]
            primary_headers = sheets_headers[primary_name]
            set_primary = set(primary_headers)

            # 构建分组信息（主表的全部列）
            grouped_columns: Dict[str, List[str]] = {}
            grouped_columns[primary_name] = list(primary_headers)

            # --- 合并信息汇总表（取金额最大的明细行的所有独有列）---
            summary_name = None
            for name in sheet_names:
                if name != primary_name and name != "建筑服务":
                    if self.GOODS_NAME_COL in sheets_headers.get(name, []):
                        summary_name = name
                        break

            summary_unique_cols: List[str] = []
            merged_summary: Dict[str, Dict[str, str]] = {}

            if summary_name and summary_name in sheets_data:
                summary_headers = sheets_headers[summary_name]
                # 找出信息汇总表独有的列（排除与主表重复的）
                summary_unique_cols = [h for h in summary_headers if h not in set_primary]
                if summary_unique_cols:
                    grouped_columns[summary_name] = summary_unique_cols

                    # 按发票号码分组
                    groups: Dict[str, List[Dict[str, Any]]] = {}
                    for row in sheets_data[summary_name]:
                        inv_num = row.get(self.INVOICE_NUMBER_COL, "")
                        if not inv_num:
                            continue
                        groups.setdefault(inv_num, []).append(row)

                    # 每组取金额最大的行
                    for inv_num, rows in groups.items():
                        best_row = rows[0]  # 默认取第一条
                        max_amount = float('-inf')
                        for row in rows:
                            try:
                                amount = float(row.get(self.AMOUNT_COL, "0") or "0")
                            except (ValueError, TypeError):
                                amount = 0
                            if amount > max_amount:
                                max_amount = amount
                                best_row = row
                        # 提取独有列的值
                        col_values = {}
                        for col in summary_unique_cols:
                            val = best_row.get(col, "")
                            # 货物或应税劳务名称列：提取商品名（去分类前缀）
                            if col == self.GOODS_NAME_COL:
                                val = self._extract_goods_name(val)
                            col_values[col] = val
                        merged_summary[inv_num] = col_values

            # --- 合并建筑服务（取独有列）---
            building_name = "建筑服务" if "建筑服务" in sheet_names else None
            building_unique_cols: List[str] = []
            merged_building: Dict[str, Dict[str, str]] = {}

            if building_name and building_name in sheets_data:
                building_headers = sheets_headers[building_name]
                # 找出建筑服务独有的列（排除与主表重复的）
                building_unique_cols = [h for h in building_headers if h not in set_primary]
                if building_unique_cols:
                    grouped_columns[building_name] = building_unique_cols
                    # 按发票号码索引
                    for row in sheets_data[building_name]:
                        inv_num = row.get(self.INVOICE_NUMBER_COL, "")
                        if not inv_num or inv_num in merged_building:
                            continue
                        merged_building[inv_num] = {col: row.get(col, "") for col in building_unique_cols}

            # --- 执行合并 ---
            merged_data = []
            for row in primary_data:
                inv_num = row.get(self.INVOICE_NUMBER_COL, "")
                merged_row = dict(row)
                # 添加信息汇总表独有列（确保所有列都存在，即使为空）
                for col in summary_unique_cols:
                    merged_row[col] = merged_summary.get(inv_num, {}).get(col, "")
                # 添加建筑服务独有列（确保所有列都存在，即使为空）
                for col in building_unique_cols:
                    merged_row[col] = merged_building.get(inv_num, {}).get(col, "")
                merged_data.append(merged_row)

            return merged_data, grouped_columns

        except ValueError:
            raise
        except Exception as e:
            raise ValueError(f"读取Excel文件失败: {str(e)}")

    def _read_sheet_headers(self, worksheet) -> List[str]:
        """读取 sheet 的表头行"""
        headers = []
        for cell in worksheet[1]:
            if cell.value is not None:
                headers.append(str(cell.value))
            else:
                headers.append(f"Column_{cell.column}")
        return headers

    def _read_sheet_data(self, worksheet, headers: List[str]) -> List[Dict[str, Any]]:
        """读取 sheet 的数据行"""
        data = []
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            row_dict = {}
            for col_idx, value in enumerate(row):
                if col_idx < len(headers):
                    if value is None:
                        row_dict[headers[col_idx]] = ""
                    else:
                        row_dict[headers[col_idx]] = str(value)
            data.append(row_dict)
        return data

    @staticmethod
    def _extract_goods_name(raw_name: str) -> str:
        """
        从"货物或应税劳务名称"中智能提取商品名或分类。

        格式: *税收分类*商品名称
        规则: 只在第一个 * 处分割，然后判断：
            - 后半部分 ≤8字 或 与分类重复 → 用后半部分(商品名)
            - 否则 → 用前半部分(分类)

        示例:
            "*黑色金属冶炼压延品*镀锌管" -> "镀锌管" (短商品名)
            "*运输服务*客运服务费" -> "客运服务费" (短商品名)
            "*日用杂品*IVSO滋味挤压油壶家用带防尘盖调料瓶" -> "日用杂品" (长商品名，用分类)
            "*乳制品*蒙纯蒙古老酸奶罐装170g*12" -> "乳制品" (长商品名，修复含*号bug)
            "*服装*服装" -> "服装" (重复，用商品名)
            "普通商品" -> "普通商品"
            "" -> ""
        """
        if not raw_name:
            return ""

        s = raw_name.strip()

        # 只在第一个 * 处分割（修复商品名含*号时取错的bug）
        first_star = s.find('*')
        if first_star == -1:
            # 没有星号，直接返回原值
            return s

        # 找第二个 * 作为分类和商品名的分界
        second_star = s.find('*', first_star + 1)

        if second_star == -1:
            # 只有一个 *，如 "*分类商品名"（少见）
            return s[first_star + 1:].strip()

        # 提取分类(第一个*和第二个*之间)和商品名(第二个*之后)
        part1 = s[first_star + 1:second_star].strip()  # 分类
        part2 = s[second_star + 1:].strip()  # 商品名

        if not part1:
            return part2 if part2 else s

        if not part2:
            return part1

        # 智能判断: 后半部分 ≤8字 或 与分类重复 → 用后半部分; 否则 → 用前半部分(分类)
        if len(part2) <= 8 or part2 == part1:
            return part2
        else:
            return part1