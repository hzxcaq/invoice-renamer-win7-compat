"""Excel读取功能测试"""
import os
import tempfile
import pytest
from openpyxl import Workbook
from ..core.excel_reader import ExcelReader


class TestExcelReader:
    """Excel读取功能测试类"""
    
    def setup_method(self):
        """测试前准备"""
        self.reader = ExcelReader()
        
    def test_read_excel_file(self):
        """测试读取Excel文件"""
        # 创建临时Excel文件
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            tmp_path = tmp.name
            
        try:
            # 创建测试数据
            wb = Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            
            # 添加表头
            headers = ["发票号码", "日期", "金额", "公司名称"]
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # 添加数据行
            test_data = [
                ["12345678", "2026-04-01", "1000.00", "测试公司A"],
                ["87654321", "2026-04-02", "2000.00", "测试公司B"],
            ]
            for row_idx, row_data in enumerate(test_data, 2):
                for col_idx, value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
            
            wb.save(tmp_path)
            
            # 读取Excel文件
            result = self.reader.read_excel(tmp_path)
            
            # 验证结果
            assert result is not None
            assert len(result) == 2
            assert result[0]["发票号码"] == "12345678"
            assert result[0]["日期"] == "2026-04-01"
            assert result[0]["金额"] == "1000.00"
            assert result[0]["公司名称"] == "测试公司A"
            
        finally:
            # 清理临时文件
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)
    
    def test_get_headers(self):
        """测试获取表头"""
        # 创建临时Excel文件
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            tmp_path = tmp.name
            
        try:
            # 创建测试数据
            wb = Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            
            # 添加表头
            headers = ["发票号码", "日期", "金额", "公司名称"]
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            wb.save(tmp_path)
            
            # 获取表头
            result = self.reader.get_headers(tmp_path)
            
            # 验证结果
            assert result is not None
            assert len(result) == 4
            assert result == headers
            
        finally:
            # 清理临时文件
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)
    
    def test_read_nonexistent_file(self):
        """测试读取不存在的文件"""
        with pytest.raises(FileNotFoundError):
            self.reader.read_excel("nonexistent.xlsx")

    def _create_multi_sheet_excel(self, tmp_path: str):
        """创建包含多个sheet的测试Excel文件"""
        wb = Workbook()
        
        # Sheet 1: 发票基础信息
        ws1 = wb.active
        ws1.title = "发票基础信息"
        headers1 = ["序号", "数电发票号码", "销方名称", "金额", "价税合计"]
        for col, h in enumerate(headers1, 1):
            ws1.cell(row=1, column=col, value=h)
        data1 = [
            [1, "INV001", "公司A", "1000", "1130"],
            [2, "INV002", "公司B", "2000", "2260"],
            [3, "INV003", "公司C", "500", "565"],
        ]
        for r, row in enumerate(data1, 2):
            for c, val in enumerate(row, 1):
                ws1.cell(row=r, column=c, value=val)
        
        # Sheet 2: 信息汇总表（INV001有2条明细，INV002有1条，INV003无）
        ws2 = wb.create_sheet("信息汇总表")
        headers2 = ["序号", "数电发票号码", "销方名称", "货物或应税劳务名称", "金额", "规格型号", "单位"]
        for col, h in enumerate(headers2, 1):
            ws2.cell(row=1, column=col, value=h)
        data2 = [
            [1, "INV001", "公司A", "*分类1*商品甲", "600", "A型", "个"],
            [2, "INV001", "公司A", "*分类2*商品乙", "400", "B型", "箱"],
            [3, "INV002", "公司B", "*分类3*商品丙", "2000", "C型", "吨"],
        ]
        for r, row in enumerate(data2, 2):
            for c, val in enumerate(row, 1):
                ws2.cell(row=r, column=c, value=val)
        
        # Sheet 3: 建筑服务（只有INV001）
        ws3 = wb.create_sheet("建筑服务")
        headers3 = ["序号", "数电发票号码", "销方名称", "建筑服务发生地", "建筑项目名称"]
        for col, h in enumerate(headers3, 1):
            ws3.cell(row=1, column=col, value=h)
        data3 = [
            [1, "INV001", "公司A", "西安市", "项目X"],
        ]
        for r, row in enumerate(data3, 2):
            for c, val in enumerate(row, 1):
                ws3.cell(row=r, column=c, value=val)
        
        wb.save(tmp_path)

    def test_read_excel_merged_basic(self):
        """测试多sheet合并基本功能"""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            tmp_path = tmp.name
        try:
            self._create_multi_sheet_excel(tmp_path)
            merged_data, grouped = self.reader.read_excel_merged(tmp_path)
            
            # 应该有3行（3个唯一发票）
            assert len(merged_data) == 3
            
            # 分组应该有3个sheet
            assert len(grouped) == 3
            assert "发票基础信息" in grouped
            assert "信息汇总表" in grouped
            assert "建筑服务" in grouped
        finally:
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)

    def test_read_excel_merged_summary_columns(self):
        """测试信息汇总表独有列合并（使用原始列名）"""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            tmp_path = tmp.name
        try:
            self._create_multi_sheet_excel(tmp_path)
            merged_data, grouped = self.reader.read_excel_merged(tmp_path)
            
            # 信息汇总表应显示3个独有列：货物或应税劳务名称、规格型号、单位
            assert "信息汇总表" in grouped
            assert "货物或应税劳务名称" in grouped["信息汇总表"]
            assert "规格型号" in grouped["信息汇总表"]
            assert "单位" in grouped["信息汇总表"]
            
            # INV001 有两条明细: 商品甲(600,A型,个) 和 商品乙(400,B型,箱)，应取商品甲
            inv1 = next(r for r in merged_data if r["数电发票号码"] == "INV001")
            assert inv1["货物或应税劳务名称"] == "商品甲"
            assert inv1["规格型号"] == "A型"
            assert inv1["单位"] == "个"
            
            # INV002 只有一条: 商品丙
            inv2 = next(r for r in merged_data if r["数电发票号码"] == "INV002")
            assert inv2["货物或应税劳务名称"] == "商品丙"
            
            # INV003 在信息汇总表中无数据，应为空
            inv3 = next(r for r in merged_data if r["数电发票号码"] == "INV003")
            assert inv3["货物或应税劳务名称"] == ""
        finally:
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)

    def test_read_excel_merged_building(self):
        """测试建筑服务列合并"""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            tmp_path = tmp.name
        try:
            self._create_multi_sheet_excel(tmp_path)
            merged_data, _ = self.reader.read_excel_merged(tmp_path)
            
            # INV001 有建筑服务信息
            inv1 = next(r for r in merged_data if r["数电发票号码"] == "INV001")
            assert inv1["建筑服务发生地"] == "西安市"
            assert inv1["建筑项目名称"] == "项目X"
            
            # INV002 无建筑服务信息
            inv2 = next(r for r in merged_data if r["数电发票号码"] == "INV002")
            assert inv2.get("建筑服务发生地", "") == ""
        finally:
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)

    def test_extract_goods_name(self):
        """测试商品名称智能提取静态方法"""
        # 短商品名(≤8字) → 用后半部分(商品名)
        assert ExcelReader._extract_goods_name("*黑色金属冶炼压延品*镀锌管") == "镀锌管"
        assert ExcelReader._extract_goods_name("*酒*酒汾酒") == "酒汾酒"
        assert ExcelReader._extract_goods_name("*餐饮服务*餐饮费") == "餐饮费"
        assert ExcelReader._extract_goods_name("*运输服务*客运服务费") == "客运服务费"
        assert ExcelReader._extract_goods_name("*日用杂品*马桶垫脚凳") == "马桶垫脚凳"

        # 长商品名(>8字) → 用前半部分(分类)
        assert ExcelReader._extract_goods_name("*日用杂品*IVSO滋味挤压油壶家用带防尘盖调料瓶厨房酱油瓶番茄酱沙拉酱瓶") == "日用杂品"
        assert ExcelReader._extract_goods_name("*乳制品*悦鲜活0乳糖鲜牛奶260ml") == "乳制品"
        assert ExcelReader._extract_goods_name("*服装*其他服装") == "其他服装"

        # 修复bug: 商品名含*号时只在第一个*处分割
        assert ExcelReader._extract_goods_name("*乳制品*蒙纯蒙古老酸奶罐装170g*12") == "乳制品"
        assert ExcelReader._extract_goods_name("*方便食品*【厂家直销】陕北特产豌豆杂粮面叶150g*10袋") == "方便食品"
        assert ExcelReader._extract_goods_name("*酒*俏饮50ml*1") == "俏饮50ml*1"  # ≤8字，用商品名

        # 分类=商品名 → 用后半部分
        assert ExcelReader._extract_goods_name("*服装*服装") == "服装"
        assert ExcelReader._extract_goods_name("*酒*酒") == "酒"
        assert ExcelReader._extract_goods_name("*餐饮服务*餐饮服务") == "餐饮服务"

        # 无星号 → 原样返回
        assert ExcelReader._extract_goods_name("普通商品") == "普通商品"
        assert ExcelReader._extract_goods_name("") == ""