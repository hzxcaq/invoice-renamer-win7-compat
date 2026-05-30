"""创建测试数据"""
import os
import tempfile
from openpyxl import Workbook


def create_test_excel(file_path):
    """创建测试Excel文件"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    
    # 添加表头
    headers = ["序号", "发票号码", "日期", "金额", "公司名称", "货物或应税劳务名称", "销方名称"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # 添加测试数据
    test_data = [
        ["1", "12345678", "2026-04-01", "1000.00", "测试公司A", "服务费", "供应商A"],
        ["2", "87654321", "2026-04-02", "2000.00", "测试公司B", "软件费", "供应商B"],
        ["3", "11111111", "2026-04-03", "3000.00", "测试公司C", "硬件费", "供应商C"],
    ]
    
    for row_idx, row_data in enumerate(test_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    wb.save(file_path)
    print(f"测试Excel文件已创建: {file_path}")


def create_test_pdf_files(directory):
    """创建测试PDF文件"""
    test_files = [
        "DZFP12345678.pdf",
        "DZFP87654321.pdf",
        "DZFP_11111111.pdf",
        "invoice_99999999.pdf",
        "other_file.pdf"
    ]
    
    for file_name in test_files:
        file_path = os.path.join(directory, file_name)
        with open(file_path, 'w') as f:
            f.write(f"这是测试PDF文件: {file_name}")
    
    print(f"测试PDF文件已创建: {len(test_files)} 个文件")


def create_test_directory():
    """创建测试目录"""
    # 创建临时目录
    test_dir = tempfile.mkdtemp(prefix="invoice_test_")
    
    # 创建Excel文件
    excel_file = os.path.join(test_dir, "2026.04取得发票.xlsx")
    create_test_excel(excel_file)
    
    # 创建PDF文件
    create_test_pdf_files(test_dir)
    
    print(f"测试目录已创建: {test_dir}")
    print(f"Excel文件: {excel_file}")
    
    return test_dir, excel_file


if __name__ == "__main__":
    test_dir, excel_file = create_test_directory()
    print("\n测试数据创建完成！")
    print(f"测试目录: {test_dir}")
    print(f"Excel文件: {excel_file}")
    print("\n您可以使用以下命令测试程序:")
    print(f"cd src/invoice_renamer")
    print(f"python main.py")
    print(f"\n然后选择目录: {test_dir}")
    print(f"选择Excel文件: {excel_file}")