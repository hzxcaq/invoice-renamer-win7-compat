"""完整流程测试"""
import os
import sys
import tempfile
import shutil

# 添加项目路径到sys.path
current_dir = os.path.dirname(os.path.abspath(__file__))
src_dir = os.path.join(current_dir, "src")
sys.path.insert(0, src_dir)

from invoice_renamer.core.excel_reader import ExcelReader
from invoice_renamer.core.file_matcher import FileMatcher
from invoice_renamer.core.renamer import Renamer


def test_full_flow():
    """测试完整流程"""
    print("开始完整流程测试...")
    
    # 创建测试目录
    test_dir = tempfile.mkdtemp(prefix="invoice_test_flow_")
    print(f"创建测试目录: {test_dir}")
    
    try:
        # 1. 创建测试数据
        print("\n1. 创建测试数据...")
        
        # 创建Excel文件
        from openpyxl import Workbook
        excel_file = os.path.join(test_dir, "test_excel.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        
        # 添加表头
        headers = ["发票号码", "金额", "公司名称"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # 添加数据
        test_data = [
            ["12345678", "1000.00", "测试公司A"],
            ["87654321", "2000.00", "测试公司B"],
        ]
        
        for row_idx, row_data in enumerate(test_data, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        wb.save(excel_file)
        print(f"  Excel文件已创建: {excel_file}")
        
        # 创建PDF文件
        pdf_files = ["DZFP12345678.pdf", "DZFP87654321.pdf", "other.pdf"]
        for pdf_name in pdf_files:
            pdf_path = os.path.join(test_dir, pdf_name)
            with open(pdf_path, 'w') as f:
                f.write(f"测试PDF: {pdf_name}")
        print(f"  PDF文件已创建: {len(pdf_files)} 个文件")
        
        # 2. 测试Excel读取
        print("\n2. 测试Excel读取...")
        excel_reader = ExcelReader()
        excel_data = excel_reader.read_excel(excel_file)
        excel_headers = excel_reader.get_headers(excel_file)
        
        print(f"  表头: {excel_headers}")
        print(f"  数据行数: {len(excel_data)}")
        for i, row in enumerate(excel_data):
            print(f"  第{i+1}行: {row}")
        
        # 3. 测试文件匹配
        print("\n3. 测试文件匹配...")
        file_matcher = FileMatcher()
        matched_files = file_matcher.match_files_with_excel(test_dir, excel_data)
        
        print(f"  匹配结果: {len(matched_files)} 个文件")
        for file_path, excel_row in matched_files:
            file_name = os.path.basename(file_path)
            print(f"  {file_name} -> {excel_row}")
        
        # 4. 测试重命名格式
        print("\n4. 测试重命名格式...")
        renamer = Renamer()
        format_template = "{公司名称}_{金额}_{发票号码}.pdf"
        
        print(f"  格式模板: {format_template}")
        
        # 生成新文件名
        rename_mapping = []
        for file_path, excel_row in matched_files:
            new_name = renamer.format_filename(format_template, excel_row)
            rename_mapping.append((file_path, new_name))
            print(f"  {os.path.basename(file_path)} -> {new_name}")
        
        # 5. 执行重命名（预览模式）
        print("\n5. 重命名预览（不实际执行）...")
        print("  如果执行重命名，文件将被重命名为:")
        for original_path, new_name in rename_mapping:
            print(f"    {os.path.basename(original_path)} -> {new_name}")
        
        # 6. 测试备份功能
        print("\n6. 测试备份功能...")
        backup_file = os.path.join(test_dir, "backup.txt")
        success = renamer.backup_original_names(rename_mapping, backup_file)
        print(f"  备份结果: {'成功' if success else '失败'}")
        
        if success:
            print(f"  备份文件: {backup_file}")
            with open(backup_file, 'r', encoding='utf-8') as f:
                content = f.read()
                print(f"  备份内容:\n{content}")
        
        print("\n完整流程测试完成！")
        return True
        
    except Exception as e:
        print(f"测试失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return False
        
    finally:
        # 清理测试目录
        print(f"\n清理测试目录: {test_dir}")
        shutil.rmtree(test_dir, ignore_errors=True)


if __name__ == "__main__":
    success = test_full_flow()
    sys.exit(0 if success else 1)